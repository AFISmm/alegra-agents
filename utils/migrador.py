"""
Clase Migrador: lógica de negocio para migrar Siigo → Alegra.
Extraída de migrador_alegra_app.py (sin dependencias de Tkinter).
"""
from __future__ import annotations

import json
import base64
import re
import time
from collections import Counter
from pathlib import Path
from datetime import datetime, date

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE_URL = "https://api.alegra.com/api/v1"

PUC_TYPE = {
    "1": "asset", "2": "liability", "3": "equity",
    "4": "income", "5": "expense", "6": "cost",
    "7": "productionCost", "8": "order", "9": "order",
}
PUC_NATURE = {
    "1": "debit",  "2": "credit", "3": "credit",
    "4": "credit", "5": "debit",  "6": "debit",
    "7": "debit",  "8": "debit",  "9": "credit",
}

REQUEST_DELAY = 0.1
RETRY_MAX     = 3
RETRY_DELAY   = 3

ALREADY_EXISTS_CODES = {32006, 32021}
JOURNAL_EXISTS_CODES = {32021, 32006}


class Migrador:
    def __init__(self, email, token, carpeta, years, log_fn, progress_fn, stop_event):
        self.email  = email
        self.token  = token
        self.folder = Path(carpeta)
        self.years  = sorted(years)
        self._log   = log_fn
        self._prog  = progress_fn
        self.stop   = stop_event

        self.headers = {
            "Authorization": "Basic " + base64.b64encode(
                f"{email}:{token}".encode()).decode(),
            "Content-Type": "application/json",
            "Accept":       "application/json",
        }

        self.accounts_cache = self.folder / "fase4_cache_cuentas.json"
        self.contacts_cache = self.folder / "fase4_cache_contactos.json"
        self.log_file       = self.folder / "fase4_log_alegra.jsonl"

        self.account_map: dict = {}
        self.contact_map: dict = {}

        self._warned_accounts: set = set()
        self._unresolvable: set = set()

        self.failed:   list[dict] = []
        self.existing: list[dict] = []
        self._accounts_reloaded = False

    # ── HTTP ──────────────────────────────────────────────────────────────────

    def _get(self, path, params=None):
        url = f"{BASE_URL}/{path.lstrip('/')}"
        for attempt in range(1, RETRY_MAX + 1):
            try:
                r = requests.get(url, params=params, headers=self.headers, timeout=30)
                r.raise_for_status()
                return r.json()
            except requests.RequestException as e:
                if attempt == RETRY_MAX:
                    raise
                self._log(f"  GET {path} intento {attempt}: {e}", "warn")
                time.sleep(RETRY_DELAY)

    def _post(self, path, payload):
        url = f"{BASE_URL}/{path.lstrip('/')}"
        for attempt in range(1, RETRY_MAX + 1):
            try:
                r = requests.post(url, json=payload, headers=self.headers, timeout=30)
                if not r.ok:
                    try:
                        body = r.json()
                    except Exception:
                        body = r.text[:300]
                    raise requests.HTTPError(f"{r.status_code}: {body}", response=r)
                return r.json()
            except requests.HTTPError:
                raise
            except requests.RequestException as e:
                if attempt == RETRY_MAX:
                    raise
                self._log(f"  POST {path} intento {attempt}: {e}", "warn")
                time.sleep(RETRY_DELAY)

    def _log_event(self, event: dict):
        try:
            with open(self.log_file, "a", encoding="utf-8") as f:
                f.write(json.dumps(event, ensure_ascii=False) + "\n")
        except Exception:
            pass

    # ── Cuentas ───────────────────────────────────────────────────────────────

    @staticmethod
    def _normalize_code(raw) -> str:
        if raw is None:
            return ""
        if isinstance(raw, float):
            return str(int(raw)) if raw == int(raw) else str(raw)
        if isinstance(raw, int):
            return str(raw)
        s = str(raw).strip()
        if re.match(r"^\d+\.0+$", s):
            return s.split(".")[0]
        return s

    def _flatten_tree(self, items):
        for item in (items or []):
            if not isinstance(item, dict):
                continue
            code = self._normalize_code(item.get("code"))
            if code and code not in self.account_map:
                self.account_map[code] = item["id"]
            children = item.get("children") or item.get("subcategories") or []
            if children:
                self._flatten_tree(children)

    def load_accounts(self, refresh=False):
        if not refresh and self.accounts_cache.exists():
            try:
                self.account_map = json.loads(
                    self.accounts_cache.read_text(encoding="utf-8"))
                self._log(f"Cache cuentas: {len(self.account_map)} entradas")
                return
            except Exception as e:
                self._log(f"Cache cuentas corrupto, recargando: {e}", "warn")
        self._fetch_all_accounts()

    def _fetch_all_accounts(self):
        self._log("Descargando catálogo completo de cuentas desde Alegra...")
        PAGE = 500
        MAX_PAGES = 30

        try:
            start = 0
            for page_num in range(MAX_PAGES):
                if self.stop.is_set():
                    break
                cats = self._get("/categories", {"start": start, "limit": PAGE, "format": "plain"})
                if not cats or not isinstance(cats, list):
                    break
                prev = len(self.account_map)
                for cat in cats:
                    if not isinstance(cat, dict):
                        continue
                    code = self._normalize_code(cat.get("code"))
                    if code:
                        self.account_map[code] = cat["id"]
                new = len(self.account_map) - prev
                self._log(f"  plain pág {page_num+1}: {len(cats)} registros, +{new} nuevas → total {len(self.account_map)}")
                if len(cats) < PAGE or new == 0:
                    break
                start += PAGE
        except Exception as e:
            self._log(f"  ERROR format=plain: {e}", "warn")

        prev_count = len(self.account_map)
        try:
            start = 0
            for page_num in range(MAX_PAGES):
                if self.stop.is_set():
                    break
                tree = self._get("/categories", {"start": start, "limit": PAGE})
                if not tree or not isinstance(tree, list):
                    break
                before = len(self.account_map)
                self._flatten_tree(tree)
                new = len(self.account_map) - before
                self._log(f"  árbol pág {page_num+1}: {len(tree)} nodos, +{new} nuevas → total {len(self.account_map)}")
                if len(tree) < PAGE or new == 0:
                    break
                start += PAGE
            extra = len(self.account_map) - prev_count
            if extra:
                self._log(f"  árbol recursivo aportó: +{extra} cuentas adicionales")
        except Exception as e:
            self._log(f"  árbol recursivo: {e}", "warn")

        self._log(f"Total cuentas: {len(self.account_map)}", "ok")
        self._save_accounts()

    @staticmethod
    def _atomic_write(path: Path, content: str):
        tmp = path.with_suffix(".tmp")
        tmp.write_text(content, encoding="utf-8")
        for attempt in range(5):
            try:
                tmp.replace(path)
                return
            except PermissionError:
                if attempt == 4:
                    try:
                        path.write_text(content, encoding="utf-8")
                    finally:
                        try:
                            tmp.unlink(missing_ok=True)
                        except Exception:
                            pass
                    return
                time.sleep(0.3 * (attempt + 1))

    def _save_accounts(self):
        self._atomic_write(self.accounts_cache, json.dumps(self.account_map, indent=2))

    def _reload_accounts_once(self):
        if self._accounts_reloaded:
            return
        self._accounts_reloaded = True
        self._log("  → Recargando catálogo completo (cuentas faltantes en cache)...", "warn")
        self._fetch_all_accounts()

    def _create_hierarchy(self, code: str, name: str) -> str | None:
        if code in self.account_map:
            return self.account_map[code]

        parent_code = code[:-2]
        if len(parent_code) < 1:
            return None

        parent_id = self._create_hierarchy(parent_code, "")
        if parent_id is None:
            return None

        clase = code[0]
        payload = {
            "name":     (name[:100] if name else f"Cuenta {code}"),
            "type":     PUC_TYPE.get(clase, "expense"),
            "nature":   PUC_NATURE.get(clase, "debit"),
            "code":     code,
            "idParent": str(parent_id),
        }
        try:
            result  = self._post("/categories", payload)
            acct_id = result["id"]
            self.account_map[code] = acct_id
            self._save_accounts()
            self._log(f"  Cuenta creada: {code}  (id={acct_id})", "ok")
            self._log_event({"ts": datetime.now().isoformat(),
                             "action": "cuenta_creada", "code": code, "id": acct_id})
            return acct_id

        except requests.HTTPError as e:
            try:
                err_body = e.response.json()
                err_code = err_body.get("code")
            except Exception:
                err_code = None

            if err_code in ALREADY_EXISTS_CODES:
                self._reload_accounts_once()
                if code in self.account_map:
                    self._log(f"  Cuenta recuperada tras recarga: {code}", "ok")
                    return self.account_map[code]
                try:
                    res = self._get("/categories", {"code": code, "limit": 5})
                    if isinstance(res, list):
                        for item in res:
                            c = self._normalize_code(item.get("code", ""))
                            if c == code:
                                acct_id = item["id"]
                                self.account_map[code] = acct_id
                                self._save_accounts()
                                self._log(f"  Cuenta encontrada por código: {code}", "ok")
                                return acct_id
                except Exception:
                    pass
                try:
                    res = self._get("/categories", {"name": f"Cuenta {code}", "limit": 5})
                    if isinstance(res, list):
                        for item in res:
                            c = self._normalize_code(item.get("code", ""))
                            if c == code:
                                acct_id = item["id"]
                                self.account_map[code] = acct_id
                                self._save_accounts()
                                self._log(f"  Cuenta encontrada por nombre: {code}", "ok")
                                return acct_id
                except Exception:
                    pass
                self._unresolvable.add(code)
                if code not in self._warned_accounts:
                    self._log(f"  ! Cuenta {code} existe en Alegra pero no se pudo localizar", "warn")
                    self._warned_accounts.add(code)
                return None

            self._log(f"  ERROR creando cuenta {code}: {e}", "err")
            return None

    def resolve_account(self, siigo_code: str, account_name: str) -> str | None:
        for length in [len(siigo_code), 8]:
            if length > len(siigo_code):
                continue
            key = siigo_code[:length]
            if key in self.account_map:
                return self.account_map[key]

        code8 = siigo_code[:8]
        if code8 in self._unresolvable:
            return None

        return self._create_hierarchy(code8, account_name)

    # ── Contactos ─────────────────────────────────────────────────────────────

    def load_contacts(self, refresh=False):
        if not refresh and self.contacts_cache.exists():
            try:
                self.contact_map = json.loads(
                    self.contacts_cache.read_text(encoding="utf-8"))
                self._log(f"Cache contactos: {len(self.contact_map)} entradas")
                return
            except Exception as e:
                self._log(f"Cache contactos corrupto, recargando: {e}", "warn")

        self._log("Descargando contactos desde Alegra...")
        nit_to_id = {}
        start = 0
        while True:
            batch = self._get("/contacts", {"start": start, "limit": 30})
            if not batch:
                break
            for c in batch:
                nit = str(c.get("identification") or "").strip()
                if nit and nit not in ("0", ""):
                    nit_to_id[nit] = str(c["id"])
            if len(batch) < 30:
                break
            start += 30
        self.contact_map = nit_to_id
        self.contacts_cache.write_text(
            json.dumps(self.contact_map, indent=2), encoding="utf-8")
        self._log(f"Contactos descargados: {len(self.contact_map)}", "ok")

    def resolve_contact(self, nit: str, nombre: str) -> str | None:
        nit = str(nit).strip().split(".")[0]
        if not nit or nit in ("0", "nan", ""):
            return None
        if nit in self.contact_map:
            return self.contact_map[nit]
        payload = {
            "name":           (nombre or f"Tercero {nit}")[:150],
            "identification": nit,
            "type":           ["client", "provider"],
        }
        try:
            result = self._post("/contacts", payload)
            cid = str(result["id"])
            self.contact_map[nit] = cid
            self.contacts_cache.write_text(
                json.dumps(self.contact_map, indent=2), encoding="utf-8")
            return cid
        except requests.HTTPError as e:
            if e.response and e.response.status_code == 400:
                try:
                    res = self._get("/contacts", {"identification": nit, "limit": 1})
                    if isinstance(res, list) and res:
                        cid = str(res[0]["id"])
                        self.contact_map[nit] = cid
                        self.contacts_cache.write_text(
                            json.dumps(self.contact_map, indent=2), encoding="utf-8")
                        return cid
                except Exception:
                    pass
            return None
        except Exception:
            return None

    # ── Parser AUXILIAR ───────────────────────────────────────────────────────

    @staticmethod
    def _parse_amount(val) -> float:
        if val is None:
            return 0.0
        if isinstance(val, (int, float)):
            return float(val)
        s = str(val).strip().replace(",", "").replace(" ", "")
        try:
            return float(s)
        except ValueError:
            return 0.0

    @staticmethod
    def _comprobante_key(ref: str) -> str:
        parts = str(ref).strip().split()
        return " ".join(parts[:3]) if len(parts) >= 3 else str(ref).strip()

    def parse_auxiliar(self, path: Path) -> list[dict]:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        ws = wb.active

        HEADER_ROW = 7
        col_idx: dict[str, int] = {}
        rows: list[dict] = []

        for row_num, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if row_num < HEADER_ROW:
                continue
            if row_num == HEADER_ROW:
                for j, cell in enumerate(row):
                    if cell is not None:
                        col_idx[str(cell).strip()] = j
                continue

            def get(name):
                j = col_idx.get(name)
                return row[j] if j is not None and j < len(row) else None

            def gets(name):
                v = get(name)
                return str(v).strip() if v is not None else ""

            raw_cuenta = get("CUENTA")
            if raw_cuenta is None:
                continue
            if isinstance(raw_cuenta, float):
                if raw_cuenta != raw_cuenta:
                    continue
                cuenta = str(int(raw_cuenta))
            elif isinstance(raw_cuenta, int):
                cuenta = str(raw_cuenta)
            else:
                cuenta = str(raw_cuenta).strip().split(".")[0]
            if not re.match(r"^\d{4,10}$", cuenta):
                continue

            raw_fecha = get("FECHA")
            if raw_fecha is None:
                continue
            if isinstance(raw_fecha, (datetime, date)):
                fecha_str = (raw_fecha.date() if isinstance(raw_fecha, datetime)
                             else raw_fecha).strftime("%Y-%m-%d")
            else:
                parsed = None
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y%m%d"):
                    try:
                        parsed = datetime.strptime(str(raw_fecha).strip(), fmt)
                        break
                    except ValueError:
                        continue
                if parsed is None:
                    continue
                fecha_str = parsed.strftime("%Y-%m-%d")

            raw_comp = get("COMPROBANTE")
            ref = str(raw_comp).strip() if raw_comp is not None else ""
            if not ref or ref in ("None", ""):
                continue

            nit     = gets("NIT").split(".")[0]
            nombre  = gets("NOMBRE")
            detalle = gets("DETALLE")[:250]
            desc    = gets("DESCRIPCION")

            j_deb  = col_idx.get("DEBITOS")
            j_cred = col_idx.get("CREDITOS")
            debito  = self._parse_amount(
                row[j_deb]  if j_deb  is not None and j_deb  < len(row) else None)
            credito = self._parse_amount(
                row[j_cred] if j_cred is not None and j_cred < len(row) else None)

            if debito == 0 and credito == 0:
                continue

            rows.append({
                "ref":     ref,
                "key":     self._comprobante_key(ref),
                "fecha":   fecha_str,
                "cuenta":  cuenta,
                "desc":    desc,
                "nit":     nit if nit not in ("0", "nan", "") else "",
                "nombre":  nombre,
                "detalle": detalle,
                "debito":  debito,
                "credito": credito,
            })

        wb.close()
        return rows

    @staticmethod
    def group_by_comprobante(rows: list[dict]) -> dict:
        groups: dict = {}
        for r in rows:
            groups.setdefault(r["key"], []).append(r)
        return groups

    # ── Checkpoint ────────────────────────────────────────────────────────────

    def _cp_path(self):
        return self.folder / "fase4_checkpoint.json"

    def load_checkpoint(self) -> dict:
        cp = self._cp_path()
        if cp.exists():
            try:
                return json.loads(cp.read_text(encoding="utf-8"))
            except Exception:
                pass
        return {}

    def save_checkpoint(self, cp: dict):
        self._atomic_write(self._cp_path(), json.dumps(cp, indent=2))

    # ── Informe de fallos ─────────────────────────────────────────────────────

    def generate_report(self) -> Path | None:
        if not self.failed and not self.existing:
            self._log("Sin fallos ni existentes — no se genera informe.", "ok")
            return None

        ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = self.folder / f"informe_fallos_{ts}.xlsx"

        wb = openpyxl.Workbook()

        hdr_font   = Font(bold=True, color="FFFFFF", size=10)
        hdr_fill   = PatternFill("solid", fgColor="1E3A5F")
        err_fill   = PatternFill("solid", fgColor="FFDACC")
        warn_fill  = PatternFill("solid", fgColor="FFF3CC")
        exist_fill = PatternFill("solid", fgColor="D6EAF8")
        center     = Alignment(horizontal="center", vertical="center")
        left       = Alignment(horizontal="left",   vertical="center", wrap_text=True)
        thin       = Side(style="thin", color="CCCCCC")
        border     = Border(left=thin, right=thin, top=thin, bottom=thin)

        def _make_hdr(ws, headers):
            for col, (hdr, width) in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=hdr)
                cell.font      = hdr_font
                cell.fill      = hdr_fill
                cell.alignment = center
                cell.border    = border
                ws.column_dimensions[
                    openpyxl.utils.get_column_letter(col)].width = width
            ws.row_dimensions[1].height = 20
            ws.freeze_panes = "A2"

        ws = wb.active
        ws.title = "Errores"
        if self.failed:
            _make_hdr(ws, [
                ("Año",         8), ("Fecha",       12), ("Comprobante", 26),
                ("Cuenta Siigo",14), ("Descripción", 40), ("Débito",      14),
                ("Crédito",     14), ("Error",       55),
            ])
            for i, rec in enumerate(self.failed, start=2):
                fill = err_fill if rec.get("error") else warn_fill
                vals = [
                    rec.get("año", ""), rec.get("fecha", ""), rec.get("comprobante", ""),
                    rec.get("cuenta", ""), rec.get("descripcion", ""),
                    rec.get("debito", 0) or "", rec.get("credito", 0) or "",
                    rec.get("error", "Sin entradas válidas"),
                ]
                for col, val in enumerate(vals, 1):
                    cell = ws.cell(row=i, column=col, value=val)
                    cell.fill      = fill
                    cell.border    = border
                    cell.alignment = center if col in (1, 2, 4, 6, 7) else left
        else:
            ws.cell(row=1, column=1, value="Sin errores en esta migración ✓")

        ws_ex = wb.create_sheet("Existentes en Alegra")
        if self.existing:
            _make_hdr(ws_ex, [
                ("Año", 8), ("Fecha", 12), ("Comprobante", 30),
                ("Descripción", 50), ("Líneas", 8),
            ])
            for i, rec in enumerate(self.existing, start=2):
                vals = [
                    rec.get("año", ""), rec.get("fecha", ""), rec.get("comprobante", ""),
                    rec.get("descripcion", ""), rec.get("lineas", ""),
                ]
                for col, val in enumerate(vals, 1):
                    cell = ws_ex.cell(row=i, column=col, value=val)
                    cell.fill      = exist_fill
                    cell.border    = border
                    cell.alignment = center if col in (1, 2, 5) else left
        else:
            ws_ex.cell(row=1, column=1,
                       value="Ningún comprobante fue detectado como ya existente.")

        ws2 = wb.create_sheet("Resumen por año")
        _make_hdr(ws2, [
            ("Año", 8), ("Subidos OK", 12), ("Existentes", 12),
            ("Errores", 10), ("Checkpoint", 12), ("% completado", 14),
        ])

        fallos_año = Counter(r.get("año", "") for r in self.failed)
        exist_año  = Counter(r.get("año", "") for r in self.existing)
        cp = self.load_checkpoint()
        for year in self.years:
            año_str = str(year)
            subidos = len(cp.get(año_str, []))
            fallos  = fallos_año.get(año_str, 0)
            existe  = exist_año.get(año_str, 0)
            total   = subidos + fallos
            pct     = f"{subidos/total*100:.1f}%" if total else "—"
            ws2.append([año_str, subidos, existe, fallos, subidos, pct])

        wb.save(out_path)
        self._log(
            f"Informe generado: {out_path.name}  "
            f"({len(self.failed)} errores, {len(self.existing)} existentes)", "ok")
        return out_path

    # ── Pipeline principal ────────────────────────────────────────────────────

    def run(self) -> Path | None:
        self._log("Cargando mapas de cuentas y contactos...")
        self.load_accounts()
        self.load_contacts()
        checkpoint = self.load_checkpoint()

        total_ok = total_err = total_skip = total_existing = 0
        total_pending_at_start = 0

        for year in self.years:
            if self.stop.is_set():
                self._log("Migración detenida por usuario.", "warn")
                break

            self._log(f"\n{'═'*52}", "hdr")
            self._log(f"  AÑO {year}", "hdr")
            self._log(f"{'═'*52}", "hdr")

            auxiliares = sorted(self.folder.glob(f"AUXILIAR *{year}*.xlsx"))
            if not auxiliares:
                self._log(f"  ✗ No se encontró AUXILIAR {year}.xlsx", "err")
                continue

            xlsx = auxiliares[0]
            self._log(f"  Archivo: {xlsx.name}")

            try:
                self._log("  Parseando comprobantes...")
                rows   = self.parse_auxiliar(xlsx)
                groups = self.group_by_comprobante(rows)
                self._log(f"  {len(rows):,} líneas → {len(groups):,} comprobantes")
            except Exception as e:
                self._log(f"  ERROR parseando {xlsx.name}: {e}", "err")
                continue

            año_str   = str(year)
            done_keys = set(checkpoint.get(año_str, []))
            self._log(f"  Ya subidos (checkpoint): {len(done_keys)}")

            ok_año = err_año = existing_año = 0
            skip_año      = len(done_keys)
            total         = len(groups)
            pending_total = total - skip_año
            proc          = 0
            total_pending_at_start += pending_total

            for comp_key, lines in sorted(groups.items()):
                if self.stop.is_set():
                    break
                if comp_key in done_keys:
                    continue

                proc += 1
                fecha   = lines[0]["fecha"]
                detalle = lines[0]["detalle"] or lines[0]["desc"]

                entries       = []
                cuentas_malas = []
                abort         = False

                for line in lines:
                    acct_id = self.resolve_account(line["cuenta"], line["desc"])
                    if acct_id is None:
                        cuentas_malas.append(line["cuenta"])
                        abort = True
                        continue

                    entry = {"id": acct_id}
                    if line["debito"] > 0:
                        entry["debit"]  = round(line["debito"],  2)
                    else:
                        entry["credit"] = round(line["credito"], 2)
                    if line["detalle"]:
                        entry["description"] = line["detalle"][:250]
                    if line["nit"]:
                        cid = self.resolve_contact(line["nit"], line["nombre"])
                        if cid:
                            entry["client"] = {"id": cid}
                    entries.append(entry)

                if abort:
                    msg = (f"Comprobante omitido — "
                           f"{len(cuentas_malas)} cuenta(s) no resueltas: "
                           f"{', '.join(cuentas_malas[:4])}")
                    self._log_event({
                        "ts": datetime.now().isoformat(), "action": "journal_omitido",
                        "año": año_str, "comp": comp_key,
                        "cuentas_malas": cuentas_malas, "error": msg,
                    })
                    err_año += 1
                    for line in lines:
                        self.failed.append({
                            "año": año_str, "fecha": fecha, "comprobante": comp_key,
                            "cuenta": line["cuenta"],
                            "descripcion": line["detalle"] or line["desc"],
                            "debito": line["debito"], "credito": line["credito"],
                            "error": (f"Cuenta no resuelta: {line['cuenta']}"
                                      if line["cuenta"] in cuentas_malas
                                      else "Comprobante abortado por otras líneas"),
                        })
                    self._prog(year, proc, pending_total, ok_año, err_año, existing_año)
                    continue

                if not entries:
                    self._log(f"  ✗ {comp_key}: sin entradas", "err")
                    err_año += 1
                    self._prog(year, proc, pending_total, ok_año, err_año, existing_año)
                    continue

                payload = {
                    "date":         fecha,
                    "reference":    comp_key[:255],
                    "observations": detalle[:500] if detalle else "",
                    "entries":      entries,
                }

                try:
                    result = self._post("/journals", payload)
                    jid    = result.get("id", "?")
                    self._log(
                        f"  ✓ {comp_key}  {fecha}  {len(entries)} líneas  id={jid}", "ok")
                    self._log_event({
                        "ts": datetime.now().isoformat(), "action": "journal_ok",
                        "año": año_str, "comp": comp_key, "fecha": fecha,
                        "id": jid, "lineas": len(entries),
                    })
                    ok_año += 1
                    done_keys.add(comp_key)
                    checkpoint[año_str] = list(done_keys)
                    self.save_checkpoint(checkpoint)

                except requests.HTTPError as e:
                    err_code = None
                    if e.response is not None:
                        try:
                            err_code = e.response.json().get("code")
                        except Exception:
                            pass
                    if err_code is None:
                        import re as _re
                        m = _re.search(r'"code"\s*:\s*(\d+)', str(e))
                        if m:
                            err_code = int(m.group(1))

                    if err_code in JOURNAL_EXISTS_CODES or (
                        e.response is not None and
                        any(kw in (e.response.text or "").lower()
                            for kw in ("ya existe", "already exists", "número ya"))
                    ):
                        self._log(f"  ↺ {comp_key}: ya existe en Alegra", "warn")
                        self._log_event({
                            "ts": datetime.now().isoformat(), "action": "journal_exists",
                            "año": año_str, "comp": comp_key, "fecha": fecha,
                        })
                        existing_año += 1
                        done_keys.add(comp_key)
                        checkpoint[año_str] = list(done_keys)
                        self.save_checkpoint(checkpoint)
                        self.existing.append({
                            "año": año_str, "fecha": fecha, "comprobante": comp_key,
                            "descripcion": detalle or "", "lineas": len(entries),
                        })
                    else:
                        err_str = str(e)[:250]
                        self._log(f"  ✗ {comp_key}: {err_str}", "err")
                        self._log_event({
                            "ts": datetime.now().isoformat(), "action": "journal_error",
                            "año": año_str, "comp": comp_key, "error": err_str,
                        })
                        err_año += 1
                        for line in lines:
                            self.failed.append({
                                "año": año_str, "fecha": fecha, "comprobante": comp_key,
                                "cuenta": line["cuenta"],
                                "descripcion": line["detalle"] or line["desc"],
                                "debito": line["debito"], "credito": line["credito"],
                                "error": err_str,
                            })

                except Exception as e:
                    err_str = str(e)[:250]
                    self._log(f"  ✗ {comp_key}: {err_str}", "err")
                    self._log_event({
                        "ts": datetime.now().isoformat(), "action": "journal_error",
                        "año": año_str, "comp": comp_key, "error": err_str,
                    })
                    err_año += 1
                    for line in lines:
                        self.failed.append({
                            "año": año_str, "fecha": fecha, "comprobante": comp_key,
                            "cuenta": line["cuenta"],
                            "descripcion": line["detalle"] or line["desc"],
                            "debito": line["debito"], "credito": line["credito"],
                            "error": err_str,
                        })

                self._prog(year, proc, pending_total, ok_año, err_año, existing_año)
                time.sleep(REQUEST_DELAY)

            total_ok       += ok_año
            total_err      += err_año
            total_skip     += skip_año
            total_existing += existing_año

            self._log(
                f"\n  Año {year} completado:  "
                f"✓ {ok_año} subidos   "
                f"↺ {existing_año} existentes   "
                f"✗ {err_año} errores   "
                f"⏭ {skip_año} checkpoint", "hdr"
            )

        self._log(
            f"\n{'═'*52}\n  MIGRACIÓN COMPLETA\n"
            f"  ✓ {total_ok} subidos   ↺ {total_existing} existentes   "
            f"✗ {total_err} errores   ⏭ {total_skip} checkpoint\n"
            f"{'═'*52}", "hdr"
        )

        report_path = self.generate_report()
        was_stopped = self.stop.is_set()
        pending_remaining = total_pending_at_start - (total_ok + total_existing + total_err)
        self._prog(
            None,
            1 if was_stopped else 0,
            max(pending_remaining, 0),
            total_ok, total_err, total_existing,
        )
        return report_path
